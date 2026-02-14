"""
Excel spreadsheet generation service.

Generates and manages .xlsx files containing extracted invoice data.
Uses openpyxl for Excel file creation and formatting.

PRIVACY NOTICE:
Excel files contain only structured extracted fields.
No raw invoice images or OCR text are stored in the spreadsheet.
Files are saved locally in the output/ directory only.
"""

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from models.schemas import InvoiceData, ValidationReport


OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "output")
EXCEL_FILENAME = "invoices.xlsx"

# Column headers matching the specification
HEADERS = [
    "Invoice Number",
    "Vendor Name",
    "Date",
    "Currency",
    "Subtotal Before VAT",
    "VAT Amount",
    "Total Including VAT",
    "Confidence Score",
    "Validation Notes",
]


def _ensure_output_dir():
    """Create output directory if it doesn't exist."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def _get_excel_path() -> str:
    """Get the full path to the Excel file."""
    _ensure_output_dir()
    return os.path.join(OUTPUT_DIR, EXCEL_FILENAME)


def _create_workbook() -> Workbook:
    """Create a new formatted workbook with headers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Extractions"

    # Header styling
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Set column widths
    column_widths = [18, 25, 14, 10, 20, 15, 20, 16, 35]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    return wb


def append_invoice_to_excel(
    invoice: InvoiceData,
    validation: ValidationReport,
) -> str:
    """
    Append extracted invoice data to the Excel spreadsheet.
    Creates the file if it doesn't exist.
    
    Args:
        invoice: Extracted invoice data.
        validation: Validation report with confidence score.
        
    Returns:
        Filename of the Excel file.
    """
    excel_path = _get_excel_path()

    # Load existing workbook or create new one
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = _create_workbook()
        ws = wb.active

    # Build validation notes string
    notes_parts = []
    for err in validation.errors:
        notes_parts.append(f"ERROR: {err.message}")
    for warn in validation.warnings:
        notes_parts.append(f"WARNING: {warn.message}")
    if validation.is_valid:
        notes_parts.insert(0, "VALID")
    validation_notes = "; ".join(notes_parts) if notes_parts else "No issues"

    # Append data row
    row_data = [
        invoice.invoice_number or "N/A",
        invoice.vendor_name or "N/A",
        invoice.invoice_date or "N/A",
        invoice.currency,
        invoice.subtotal,
        invoice.vat_amount,
        invoice.total,
        round(validation.confidence_score, 1),
        validation_notes,
    ]

    next_row = ws.max_row + 1
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=next_row, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")

        # Format currency columns
        if col_idx in (5, 6, 7) and isinstance(value, (int, float)):
            cell.number_format = '#,##0.00'

    # Save workbook
    wb.save(excel_path)
    wb.close()

    return EXCEL_FILENAME


def get_excel_path() -> str:
    """Get the full filesystem path to the Excel output file."""
    return _get_excel_path()
